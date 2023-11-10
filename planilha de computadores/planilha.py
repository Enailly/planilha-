import openpyxl 

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet("Computadores")

computadores_page = book["Computadores"]
computadores_page.append(["Eletrônica", "Memória ram", "preço"])

computadores_page.append(["Computador 1", "8gb ram", "R$ 2500"])
computadores_page.append(["Computador 2", "16gb ram", "R$ 5500"])
computadores_page.append(["Computador 3", "32gb ram", "R$ 8500"])

#salva a pĺanilha 
book.save("planilha de computadores.xlsx")




